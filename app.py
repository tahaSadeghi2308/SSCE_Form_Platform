# -*- coding: utf-8 -*-

import io
import functools
from datetime import datetime
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, jsonify
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import config
import bale_client

app = Flask(__name__)
app.config["SECRET_KEY"] = config.SECRET_KEY

def has_course():
    return bool(session.get("course"))

def has_professor():
    return has_course() and bool(session.get("professor"))

def has_team_members():
    return has_professor() and len(session.get("members", [])) >= 1

def has_leader():
    return has_team_members() and bool(session.get("leader_telegram_id"))

def has_submitted():
    return bool(session.get("submitted"))


def require_step(check_fn, redirect_to):
    def decorator(view_func):
        @functools.wraps(view_func)
        def wrapped(*args, **kwargs):
            if not check_fn():
                return redirect(url_for(redirect_to))
            return view_func(*args, **kwargs)
        return wrapped
    return decorator


require_course = require_step(has_course, "course")
require_professor = require_step(has_professor, "professor")
require_team_members = require_step(has_team_members, "team")
require_leader = require_step(has_leader, "leader")
require_submitted = require_step(has_submitted, "welcome")


def next_member_id():
    seq = session.get("member_seq", 0) + 1
    session["member_seq"] = seq
    return seq

@app.route("/")
def welcome():
    return render_template("welcome.html", site_title=config.SITE_TITLE)

@app.route("/course", methods=["GET", "POST"])
def course():
    error = None
    value = session.get("course", "")

    if request.method == "POST":
        value = (request.form.get("course_name") or "").strip()
        if not value:
            error = "وارد کردن نام درس الزامی است."
        else:
            session["course"] = value
            return redirect(url_for("professor"))

    return render_template(
        "course.html", value=value, error=error, step=1, total_steps=4
    )

@app.route("/professor", methods=["GET", "POST"])
@require_course
def professor():
    error = None
    value = session.get("professor", "")

    if request.method == "POST":
        value = (request.form.get("professor_name") or "").strip()
        if not value:
            error = "وارد کردن نام استاد الزامی است."
        else:
            session["professor"] = value
            return redirect(url_for("team"))

    return render_template(
        "professor.html", value=value, error=error, step=2, total_steps=4
    )

@app.route("/team")
@require_professor
def team():
    return render_template(
        "team.html",
        members=session.get("members", []),
        roles=config.TEAM_ROLES,
        course=session.get("course"),
        professor=session.get("professor"),
        step=3, total_steps=4,
    )


@app.route("/api/members", methods=["POST"])
@require_professor
def add_member():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    role = (data.get("role") or "").strip()

    if not name:
        return jsonify(ok=False, error="نام عضو نمی‌تواند خالی باشد."), 400
    if not role or role not in config.TEAM_ROLES:
        return jsonify(ok=False, error="لطفاً یک سمت معتبر انتخاب کنید."), 400

    members = session.get("members", [])
    members.append({"id": next_member_id(), "name": name, "role": role})
    session["members"] = members
    session.modified = True
    return jsonify(ok=True, members=members)


@app.route("/api/members/<int:member_id>", methods=["PUT"])
@require_professor
def edit_member(member_id):
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    role = (data.get("role") or "").strip()

    if not name:
        return jsonify(ok=False, error="نام عضو نمی‌تواند خالی باشد."), 400
    if not role or role not in config.TEAM_ROLES:
        return jsonify(ok=False, error="لطفاً یک سمت معتبر انتخاب کنید."), 400

    members = session.get("members", [])
    found = False
    for m in members:
        if m["id"] == member_id:
            m["name"] = name
            m["role"] = role
            found = True
            break

    if not found:
        return jsonify(ok=False, error="عضو مورد نظر یافت نشد."), 404

    session["members"] = members
    session.modified = True
    return jsonify(ok=True, members=members)


@app.route("/api/members/<int:member_id>", methods=["DELETE"])
@require_professor
def delete_member(member_id):
    members = session.get("members", [])
    new_members = [m for m in members if m["id"] != member_id]

    if len(new_members) == len(members):
        return jsonify(ok=False, error="عضو مورد نظر یافت نشد."), 404

    session["members"] = new_members
    session.modified = True
    return jsonify(ok=True, members=new_members)

@app.route("/leader", methods=["GET", "POST"])
@require_team_members
def leader():
    error = None
    value = session.get("leader_telegram_id", "")

    if request.method == "POST":
        value = (request.form.get("leader_telegram_id") or "").strip()
        if not value:
            error = "وارد کردن آیدی تلگرام سرپرست تیم الزامی است."
        else:
            session["leader_telegram_id"] = value
            return redirect(url_for("summary"))

    return render_template(
        "leader.html", value=value, error=error, step=4, total_steps=4
    )

@app.route("/summary")
@require_leader
def summary():
    return render_template(
        "summary.html",
        course=session.get("course"),
        professor=session.get("professor"),
        members=session.get("members", []),
        leader_telegram_id=session.get("leader_telegram_id"),
    )

def build_xlsx(course_name, professor_name, members, leader_telegram_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "ثبت TA"
    ws.sheet_view.rightToLeft = True

    bold = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16423C")
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = "نام درس"
    ws["B1"] = course_name
    ws["A2"] = "نام استاد"
    ws["B2"] = professor_name
    ws["A3"] = "آیدی تلگرام سرپرست تیم"
    ws["B3"] = leader_telegram_id
    ws["A4"] = "تاریخ ثبت"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row in (1, 2, 3, 4):
        ws[f"A{row}"].font = bold

    header_row = 6
    headers = ["ردیف", "نام عضو تیم", "سمت"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, member in enumerate(members, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).alignment = center
        ws.cell(row=r, column=2, value=member["name"])
        ws.cell(row=r, column=3, value=member["role"])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/submit", methods=["POST"])
@require_leader
def submit():
    course_name = session.get("course")
    professor_name = session.get("professor")
    members = session.get("members", [])
    leader_telegram_id = session.get("leader_telegram_id")

    if not course_name or not professor_name:
        return jsonify(ok=False, error="اطلاعات درس یا استاد ناقص است."), 400
    if len(members) < 1:
        return jsonify(ok=False, error="حداقل باید یک عضو تیم ثبت شود."), 400
    if not leader_telegram_id:
        return jsonify(ok=False, error="آیدی تلگرام سرپرست تیم ثبت نشده است."), 400

    caption = (
        f"ثبت TA\nدرس: {course_name}\nاستاد: {professor_name}\n"
        f"تعداد اعضا: {len(members)}\nآیدی تلگرام سرپرست: {leader_telegram_id}"
    )
    filename = f"ta_registration_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        xlsx_buffer = build_xlsx(course_name, professor_name, members, leader_telegram_id)
        bale_client.send_document(xlsx_buffer, filename, caption)
    except Exception as exc:
        return jsonify(ok=False, error=f"ارسال اطلاعات با خطا مواجه شد: {exc}"), 502

    session.clear()
    session["submitted"] = True
    return jsonify(ok=True)

@app.route("/thanks")
@require_submitted
def thanks():
    return render_template("thanks.html")

@app.route("/start-over", methods=["POST"])
def start_over():
    session.clear()
    return redirect(url_for("welcome"))

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=3225, threaded=True)
